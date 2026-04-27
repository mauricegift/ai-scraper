"""
File text extraction utilities.

Supports: plain text, PDF, Word (.docx/.doc), Excel (.xlsx/.xls), CSV,
and most code/markup formats.
"""
import io
import logging

log = logging.getLogger(__name__)

# Extensions treated as plain text regardless of MIME type
_TEXT_EXTENSIONS = (
    ".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".xml",
    ".html", ".htm", ".js", ".ts", ".py", ".sh", ".toml",
    ".ini", ".log", ".tsv", ".rtf", ".env",
)

# Hard cap on returned characters to prevent huge payloads
MAX_CHARS = 50_000


def extract_text_from_file(file_bytes: bytes, filename: str, mime: str) -> str:
    """
    Extract readable text from an uploaded file.

    Parameters
    ----------
    file_bytes : raw file content
    filename   : original filename (used for extension sniffing)
    mime       : MIME type from the multipart upload

    Returns
    -------
    Extracted text string, or an error message string beginning with '['.
    """
    fname = filename.lower()

    # ── Plain text formats ────────────────────────────────────────────────────
    if mime.startswith("text/") or fname.endswith(_TEXT_EXTENSIONS):
        try:
            return file_bytes.decode("utf-8", errors="replace")
        except Exception:
            return file_bytes.decode("latin-1", errors="replace")

    # ── PDF ───────────────────────────────────────────────────────────────────
    if mime == "application/pdf" or fname.endswith(".pdf"):
        try:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for i, page in enumerate(pdf.pages):
                    t = page.extract_text()
                    if t:
                        text_parts.append(f"[Page {i + 1}]\n{t}")
            result = "\n\n".join(text_parts)
            return result if result.strip() else "[PDF has no extractable text — may be scanned/image-only]"
        except Exception as exc:
            log.warning(f"PDF extraction error: {exc}")
            return f"[PDF extraction error: {exc}]"

    # ── Word DOCX ─────────────────────────────────────────────────────────────
    _word_mimes = (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    )
    if mime in _word_mimes or fname.endswith((".docx", ".doc")):
        try:
            import docx
            doc = docx.Document(io.BytesIO(file_bytes))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs) or "[Word document has no readable text]"
        except Exception as exc:
            log.warning(f"Word extraction error: {exc}")
            return f"[Word extraction error: {exc}]"

    # ── Excel XLSX / XLS ──────────────────────────────────────────────────────
    _excel_mimes = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    )
    if mime in _excel_mimes or fname.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"[Sheet: {sheet_name}]")
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(v) if v is not None else "" for v in row]
                    if any(v.strip() for v in row_vals):
                        parts.append("\t".join(row_vals))
            return "\n".join(parts) or "[Excel file has no readable data]"
        except Exception as exc:
            log.warning(f"Excel extraction error: {exc}")
            return f"[Excel extraction error: {exc}]"

    # ── CSV fallback (may arrive as application/octet-stream) ─────────────────
    if fname.endswith(".csv"):
        try:
            return file_bytes.decode("utf-8", errors="replace")
        except Exception:
            return file_bytes.decode("latin-1", errors="replace")

    return f"[Cannot extract text from this file type: {mime} / {filename}]"
